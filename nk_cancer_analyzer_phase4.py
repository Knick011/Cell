"""
NK Cell Cancer Analysis Tool - Phase 4: Time Series Analysis and NK Cell Tracking
Author: AI Assistant
Date: 2025-01-20

This module adds:
- Cancer cell tracking across timepoints
- NK cell detection and tracking
- Death event detection
- Time series analysis and Excel export
"""

import numpy as np
import pandas as pd
import cv2
from scipy import ndimage
from scipy.optimize import linear_sum_assignment
import matplotlib.pyplot as plt
from matplotlib import patches, animation
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Import our previous modules
from nk_cancer_analyzer import ND2Analyzer
from nk_cancer_analyzer_phase3 import DropletDetector, DropletCellAnalyzer

class NKCellDetector:
    """Detect NK cells in brightfield images."""
    
    def __init__(self, pixel_size_um=0.65):
        """
        Initialize NK cell detector.
        NK cells are smaller than cancer cells and appear as empty spheres.
        """
        self.pixel_size_um = pixel_size_um
        # NK cells are smaller than cancer cells (10-12 µm vs 15-20 µm)
        self.expected_diameter_um = 11  # µm
        self.expected_diameter_px = self.expected_diameter_um / pixel_size_um
        
    def detect_nk_cells(self, brightfield_image, droplet_mask=None):
        """
        Detect NK cells as small, dark-ringed circles in brightfield.
        
        Args:
            brightfield_image: Brightfield channel
            droplet_mask: Optional mask to search only within droplets
            
        Returns:
            List of NK cell positions and properties
        """
        # Preprocess image
        img_norm = cv2.normalize(brightfield_image, None, 0, 255, cv2.NORM_MINMAX)
        img_8bit = img_norm.astype(np.uint8)
        
        # Apply mask if provided
        if droplet_mask is not None:
            img_8bit = cv2.bitwise_and(img_8bit, img_8bit, mask=droplet_mask.astype(np.uint8))
        
        # Enhance edges to detect bubble-like structures
        # NK cells appear as dark rings (empty spheres)
        blurred = cv2.GaussianBlur(img_8bit, (5, 5), 1)
        
        # Use gradient magnitude to find ring-like structures
        sobelx = cv2.Sobel(blurred, cv2.CV_64F, 1, 0, ksize=3)
        sobely = cv2.Sobel(blurred, cv2.CV_64F, 0, 1, ksize=3)
        gradient_mag = np.sqrt(sobelx**2 + sobely**2)
        
        # Normalize gradient
        gradient_norm = cv2.normalize(gradient_mag, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
        
        # Detect circles (NK cells)
        min_radius = int(self.expected_diameter_px * 0.3)  # NK cells are small
        max_radius = int(self.expected_diameter_px * 0.7)
        
        circles = cv2.HoughCircles(
            gradient_norm,
            cv2.HOUGH_GRADIENT,
            dp=1,
            minDist=int(self.expected_diameter_px * 0.8),
            param1=30,
            param2=15,  # Low threshold to catch faint circles
            minRadius=min_radius,
            maxRadius=max_radius
        )
        
        nk_cells = []
        if circles is not None:
            circles = np.round(circles[0, :]).astype("int")
            
            for i, (x, y, r) in enumerate(circles):
                # Verify it's an NK cell by checking:
                # 1. Dark ring structure
                # 2. Bright center (empty)
                if self._verify_nk_cell(img_8bit, x, y, r):
                    nk_cells.append({
                        'id': i + 1,
                        'x': x,
                        'y': y,
                        'radius': r,
                        'diameter_um': 2 * r * self.pixel_size_um
                    })
        
        return nk_cells
    
    def _verify_nk_cell(self, image, cx, cy, r):
        """Verify NK cell by checking for empty sphere morphology."""
        # Create masks
        y, x = np.ogrid[:image.shape[0], :image.shape[1]]
        
        # Center region (should be bright)
        center_mask = (x - cx)**2 + (y - cy)**2 <= (r * 0.5)**2
        
        # Ring region (should be dark)
        ring_mask = ((x - cx)**2 + (y - cy)**2 <= r**2) & \
                   ((x - cx)**2 + (y - cy)**2 >= (r * 0.7)**2)
        
        if np.sum(center_mask) == 0 or np.sum(ring_mask) == 0:
            return False
        
        # Check intensities
        center_intensity = np.mean(image[center_mask])
        ring_intensity = np.mean(image[ring_mask])
        
        # NK cells have bright centers and dark rings
        return center_intensity > ring_intensity * 1.2


class CellTracker:
    """Enhanced cell tracker for time series analysis."""
    
    def __init__(self, max_distance=30, max_frames_missing=3):
        """
        Initialize tracker.
        
        Args:
            max_distance: Maximum distance (pixels) to link cells between frames
            max_frames_missing: Maximum frames a cell can be missing before considered dead
        """
        self.max_distance = max_distance
        self.max_frames_missing = max_frames_missing
        self.tracks = {}
        self.next_id = 1
        
    def update(self, detections, frame_num, detection_type='cancer'):
        """
        Update tracks with new detections.
        
        Args:
            detections: List of detected cells with 'x', 'y' coordinates
            frame_num: Current frame number
            detection_type: 'cancer' or 'nk'
            
        Returns:
            assignments: Dict mapping detection indices to track IDs
        """
        # Get active tracks
        active_tracks = {tid: track for tid, track in self.tracks.items() 
                        if track['type'] == detection_type and 
                        frame_num - track['last_seen'] <= self.max_frames_missing}
        
        if not active_tracks or not detections:
            # Assign new IDs to all detections
            assignments = {}
            for i, det in enumerate(detections):
                track_id = self.next_id
                self.next_id += 1
                
                self.tracks[track_id] = {
                    'type': detection_type,
                    'first_frame': frame_num,
                    'last_seen': frame_num,
                    'positions': [(det.get('x', det.get('centroid_x')), 
                                 det.get('y', det.get('centroid_y')))],
                    'intensities': [det.get('mean_intensity', 1.0)],
                    'status': 'active',
                    'death_frame': None
                }
                assignments[i] = track_id
            return assignments
        
        # Build cost matrix for assignment
        track_ids = list(active_tracks.keys())
        n_tracks = len(track_ids)
        n_dets = len(detections)
        
        cost_matrix = np.full((n_tracks, n_dets), self.max_distance * 2)
        
        for i, tid in enumerate(track_ids):
            track = active_tracks[tid]
            last_pos = track['positions'][-1]
            
            for j, det in enumerate(detections):
                x = det.get('x', det.get('centroid_x'))
                y = det.get('y', det.get('centroid_y'))
                
                dist = np.sqrt((x - last_pos[0])**2 + (y - last_pos[1])**2)
                if dist < self.max_distance:
                    cost_matrix[i, j] = dist
        
        # Solve assignment problem
        if n_tracks > 0 and n_dets > 0:
            row_indices, col_indices = linear_sum_assignment(cost_matrix)
        else:
            row_indices, col_indices = [], []
        
        assignments = {}
        assigned_tracks = set()
        assigned_dets = set()
        
        # Process assignments
        for row, col in zip(row_indices, col_indices):
            if cost_matrix[row, col] < self.max_distance:
                tid = track_ids[row]
                det = detections[col]
                
                # Update track
                x = det.get('x', det.get('centroid_x'))
                y = det.get('y', det.get('centroid_y'))
                
                self.tracks[tid]['positions'].append((x, y))
                self.tracks[tid]['intensities'].append(det.get('mean_intensity', 1.0))
                self.tracks[tid]['last_seen'] = frame_num
                
                assignments[col] = tid
                assigned_tracks.add(tid)
                assigned_dets.add(col)
        
        # Create new tracks for unassigned detections
        for j, det in enumerate(detections):
            if j not in assigned_dets:
                track_id = self.next_id
                self.next_id += 1
                
                x = det.get('x', det.get('centroid_x'))
                y = det.get('y', det.get('centroid_y'))
                
                self.tracks[track_id] = {
                    'type': detection_type,
                    'first_frame': frame_num,
                    'last_seen': frame_num,
                    'positions': [(x, y)],
                    'intensities': [det.get('mean_intensity', 1.0)],
                    'status': 'active',
                    'death_frame': None
                }
                assignments[j] = track_id
        
        return assignments
    
    def mark_dead_cells(self, frame_num, intensity_threshold=0.3):
        """Mark cells as dead based on intensity drop or disappearance."""
        for tid, track in self.tracks.items():
            if track['type'] == 'cancer' and track['status'] == 'active':
                # Check if cell hasn't been seen recently
                if frame_num - track['last_seen'] > self.max_frames_missing:
                    track['status'] = 'dead'
                    track['death_frame'] = track['last_seen']
                
                # Check for intensity drop
                elif len(track['intensities']) > 2:
                    current_intensity = track['intensities'][-1]
                    max_intensity = max(track['intensities'])
                    
                    if current_intensity < intensity_threshold * max_intensity:
                        track['status'] = 'dying'
                        if track.get('death_frame') is None:
                            track['death_frame'] = frame_num


class TimeSeriesAnalyzer:
    """Analyze time series data for NK killing dynamics."""
    
    def __init__(self, nd2_analyzer, time_interval_min=15):
        """
        Initialize analyzer.
        
        Args:
            nd2_analyzer: ND2Analyzer instance
            time_interval_min: Time between frames in minutes
        """
        self.analyzer = nd2_analyzer
        self.time_interval = time_interval_min
        self.results = []
        
    def analyze_full_time_series(self, output_dir=None):
        """
        Analyze all timepoints in the ND2 file.
        
        Returns:
            DataFrame with complete time series analysis
        """
        print(f"Analyzing {self.analyzer.metadata['frames']} timepoints...")
        
        # Initialize detectors
        droplet_detector = DropletDetector()
        nk_detector = NKCellDetector()
        cell_analyzer = DropletCellAnalyzer(droplet_detector, None)
        
        # Initialize trackers
        cancer_tracker = CellTracker(max_distance=30)
        nk_tracker = CellTracker(max_distance=50)  # NK cells move more
        
        # Detect droplets in first frame
        bf_frame = self.analyzer.get_frame(0, 'brightfield')
        droplets = droplet_detector.detect_droplets(bf_frame)
        masks, combined_mask = droplet_detector.create_droplet_masks(bf_frame.shape, droplets)
        
        print(f"Tracking cells in {len(droplets)} droplets...")
        
        # Process each timepoint
        frame_results = []
        
        for t in range(self.analyzer.metadata['frames']):
            if t % 10 == 0:
                print(f"Processing frame {t}/{self.analyzer.metadata['frames']-1}")
            
            # Get frames
            bf_frame = self.analyzer.get_frame(t, 'brightfield')
            tritc_frame = self.analyzer.get_frame(t, 'TRITC')
            
            if bf_frame is None or tritc_frame is None:
                continue
            
            # Detect NK cells
            nk_cells = nk_detector.detect_nk_cells(bf_frame, combined_mask)
            nk_assignments = nk_tracker.update(nk_cells, t, 'nk')
            
            # Analyze each droplet
            droplet_frame_data = []
            
            for droplet in droplets:
                droplet_id = droplet['id']
                mask = masks[droplet_id]
                
                # Extract regions
                masked_tritc = tritc_frame.copy()
                masked_tritc[~mask] = 0
                masked_bf = bf_frame.copy()
                masked_bf[~mask] = 0
                
                # Detect cancer cells using the nuclear detection method
                nuclei = cell_analyzer._detect_nuclei(masked_tritc, masked_bf, mask, droplet)
                
                # Convert to detection format
                cancer_cells = []
                for nx, ny, intensity, area in nuclei:
                    cancer_cells.append({
                        'centroid_x': nx,
                        'centroid_y': ny,
                        'mean_intensity': intensity,
                        'area': area
                    })
                
                # Update cancer cell tracks
                cancer_assignments = cancer_tracker.update(cancer_cells, t, 'cancer')
                
                # Mark dead cells
                cancer_tracker.mark_dead_cells(t)
                
                # Count NK cells in this droplet
                nk_in_droplet = 0
                for nk in nk_cells:
                    dist = np.sqrt((nk['x'] - droplet['center_x'])**2 + 
                                 (nk['y'] - droplet['center_y'])**2)
                    if dist < droplet['radius_px']:
                        nk_in_droplet += 1
                
                # Get cancer cell states
                alive_cancer = 0
                dying_cancer = 0
                dead_cancer = 0
                
                for i, cell in enumerate(cancer_cells):
                    if i in cancer_assignments:
                        track = cancer_tracker.tracks[cancer_assignments[i]]
                        if track['status'] == 'active':
                            alive_cancer += 1
                        elif track['status'] == 'dying':
                            dying_cancer += 1
                
                # Count total dead from tracks
                for tid, track in cancer_tracker.tracks.items():
                    if track['type'] == 'cancer' and track['status'] == 'dead':
                        # Check if death occurred in this droplet
                        if track['positions']:
                            last_x, last_y = track['positions'][-1]
                            dist = np.sqrt((last_x - droplet['center_x'])**2 + 
                                         (last_y - droplet['center_y'])**2)
                            if dist < droplet['radius_px']:
                                dead_cancer += 1
                
                droplet_frame_data.append({
                    'timepoint': t,
                    'time_min': t * self.time_interval,
                    'droplet_id': droplet_id,
                    'droplet_type': droplet['type'],
                    'nk_cells': nk_in_droplet,
                    'cancer_cells_alive': alive_cancer,
                    'cancer_cells_dying': dying_cancer,
                    'cancer_cells_dead': dead_cancer,
                    'total_cancer_cells': alive_cancer + dying_cancer
                })
            
            frame_results.extend(droplet_frame_data)
        
        # Create DataFrame
        df = pd.DataFrame(frame_results)
        
        # Add death timing information
        death_events = []
        for tid, track in cancer_tracker.tracks.items():
            if track['type'] == 'cancer' and track['death_frame'] is not None:
                death_events.append({
                    'cell_id': tid,
                    'death_frame': track['death_frame'],
                    'death_time_min': track['death_frame'] * self.time_interval,
                    'lifespan_frames': track['death_frame'] - track['first_frame'],
                    'lifespan_min': (track['death_frame'] - track['first_frame']) * self.time_interval
                })
        
        death_df = pd.DataFrame(death_events)
        
        # Save results
        if output_dir:
            self._save_results(df, death_df, output_dir)
        
        return df, death_df, cancer_tracker, nk_tracker
    
    def _save_results(self, time_series_df, death_events_df, output_dir):
        """Save results to Excel file with formatting."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"NK_killing_analysis_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Time series data
            time_series_df.to_excel(writer, sheet_name='Time_Series', index=False)
            
            # Death events
            if len(death_events_df) > 0:
                death_events_df.to_excel(writer, sheet_name='Death_Events', index=False)
            
            # Summary statistics
            summary_data = self._calculate_summary_stats(time_series_df, death_events_df)
            summary_df = pd.DataFrame([summary_data])
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format the Excel file
            workbook = writer.book
            
            # Format time series sheet
            ws = workbook['Time_Series']
            self._format_sheet(ws)
            
            # Format death events sheet
            if 'Death_Events' in workbook.sheetnames:
                ws = workbook['Death_Events']
                self._format_sheet(ws)
            
            # Format summary sheet
            ws = workbook['Summary']
            self._format_sheet(ws, highlight=True)
        
        print(f"Results saved to: {filepath}")
        return filepath
    
    def _format_sheet(self, worksheet, highlight=False):
        """Apply formatting to Excel worksheet."""
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Highlight important values
        if highlight:
            important_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if cell.column_letter in ['B', 'C', 'D']:  # Key metrics
                        cell.fill = important_fill
    
    def _calculate_summary_stats(self, time_series_df, death_events_df):
        """Calculate summary statistics."""
        total_droplets = time_series_df['droplet_id'].nunique()
        
        # Initial and final cell counts
        initial_cancer = time_series_df[time_series_df['timepoint'] == 0]['cancer_cells_alive'].sum()
        final_cancer = time_series_df[time_series_df['timepoint'] == time_series_df['timepoint'].max()]['cancer_cells_alive'].sum()
        
        # Death statistics
        total_deaths = len(death_events_df) if len(death_events_df) > 0 else 0
        avg_death_time = death_events_df['death_time_min'].mean() if len(death_events_df) > 0 else 'N/A'
        
        # NK cell statistics
        max_nk = time_series_df['nk_cells'].max()
        avg_nk_per_droplet = time_series_df.groupby('timepoint')['nk_cells'].sum().mean()
        
        return {
            'total_droplets': total_droplets,
            'initial_cancer_cells': initial_cancer,
            'final_cancer_cells': final_cancer,
            'total_deaths': total_deaths,
            'killing_efficiency_%': (total_deaths / initial_cancer * 100) if initial_cancer > 0 else 0,
            'average_death_time_min': avg_death_time,
            'max_nk_cells': max_nk,
            'avg_nk_per_droplet': avg_nk_per_droplet,
            'total_frames': time_series_df['timepoint'].max() + 1,
            'total_duration_min': time_series_df['time_min'].max()
        }
    
    def create_killing_visualization(self, time_series_df, output_path=None):
        """Create visualization of killing dynamics."""
        fig, axes = plt.subplots(2, 2, figsize=(15, 10))
        
        # Plot 1: Cancer cells over time by droplet
        ax = axes[0, 0]
        for droplet_id in time_series_df['droplet_id'].unique():
            droplet_data = time_series_df[time_series_df['droplet_id'] == droplet_id]
            ax.plot(droplet_data['time_min'], droplet_data['cancer_cells_alive'], 
                   label=f'Droplet {droplet_id}', marker='o', markersize=4)
        
        ax.set_xlabel('Time (minutes)')
        ax.set_ylabel('Live Cancer Cells')
        ax.set_title('Cancer Cell Survival by Droplet')
        ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
        ax.grid(True, alpha=0.3)
        
        # Plot 2: NK cells over time
        ax = axes[0, 1]
        nk_by_time = time_series_df.groupby('time_min')['nk_cells'].sum()
        ax.plot(nk_by_time.index, nk_by_time.values, 'b-', linewidth=2, marker='s')
        ax.set_xlabel('Time (minutes)')
        ax.set_ylabel('Total NK Cells')
        ax.set_title('NK Cell Count Over Time')
        ax.grid(True, alpha=0.3)
        
        # Plot 3: Cumulative deaths
        ax = axes[1, 0]
        dead_by_time = time_series_df.groupby('time_min')['cancer_cells_dead'].sum()
        ax.plot(dead_by_time.index, dead_by_time.values, 'r-', linewidth=2, marker='^')
        ax.set_xlabel('Time (minutes)')
        ax.set_ylabel('Cumulative Deaths')
        ax.set_title('Cumulative Cancer Cell Deaths')
        ax.grid(True, alpha=0.3)
        
        # Plot 4: Killing efficiency
        ax = axes[1, 1]
        total_cancer = time_series_df.groupby('time_min')['total_cancer_cells'].sum()
        initial_count = total_cancer.iloc[0] if len(total_cancer) > 0 else 1
        survival_pct = (total_cancer / initial_count) * 100
        
        ax.plot(survival_pct.index, survival_pct.values, 'g-', linewidth=2)
        ax.set_xlabel('Time (minutes)')
        ax.set_ylabel('Survival (%)')
        ax.set_title('Cancer Cell Survival Percentage')
        ax.set_ylim(0, 105)
        ax.grid(True, alpha=0.3)
        
        plt.tight_layout()
        
        if output_path:
            plt.savefig(output_path, dpi=300, bbox_inches='tight')
            print(f"Visualization saved to: {output_path}")
        
        plt.show()


# Main analysis function
def analyze_nk_killing(nd2_file, output_dir=None):
    """
    Complete analysis pipeline for NK killing.
    
    Args:
        nd2_file: Path to ND2 file
        output_dir: Directory to save results (creates one if None)
        
    Returns:
        Dictionary with analysis results
    """
    print(f"Starting NK killing analysis for: {nd2_file}")
    
    # Create output directory
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(nd2_file), "analysis_results")
    os.makedirs(output_dir, exist_ok=True)
    
    # Load ND2 file
    analyzer = ND2Analyzer(nd2_file)
    if not analyzer.load_file():
        print("Failed to load ND2 file!")
        return None
    
    # Run time series analysis
    ts_analyzer = TimeSeriesAnalyzer(analyzer)
    time_series_df, death_events_df, cancer_tracker, nk_tracker = ts_analyzer.analyze_full_time_series(output_dir)
    
    # Create visualizations
    vis_path = os.path.join(output_dir, "killing_dynamics.png")
    ts_analyzer.create_killing_visualization(time_series_df, vis_path)
    
    # Close analyzer
    analyzer.close()
    
    print("\nAnalysis complete!")
    print(f"Results saved to: {output_dir}")
    
    return {
        'time_series': time_series_df,
        'death_events': death_events_df,
        'cancer_tracks': cancer_tracker.tracks,
        'nk_tracks': nk_tracker.tracks,
        'output_dir': output_dir
    }


# Test function
def test_phase4():
    """Test Phase 4 functionality."""
    print("Phase 4: Time Series Analysis Test")
    print("=" * 50)
    
    # Your ND2 file
    nd2_file = r"D:\New\BrainBites\Cell\03.nd2"
    
    # Run analysis
    results = analyze_nk_killing(nd2_file)
    
    if results:
        print("\nAnalysis Summary:")
        print(f"Total timepoints analyzed: {results['time_series']['timepoint'].max() + 1}")
        print(f"Total cancer cell tracks: {len([t for t in results['cancer_tracks'].values() if t['type'] == 'cancer'])}")
        print(f"Total NK cell tracks: {len([t for t in results['nk_tracks'].values() if t['type'] == 'nk'])}")
        print(f"Death events detected: {len(results['death_events'])}")


if __name__ == "__main__":
    test_phase4()